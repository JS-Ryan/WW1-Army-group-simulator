import openpyxl
import random
import os
import subprocess

# Create a 6x6 blank worksheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Map"

# Set the width and height of each cell
for i in range(1, 7):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10
    sheet.row_dimensions[i].height = 50

# Define terrain colors
terrain_colors = {
    "Mountain": "8B4513",  # Brown
    "Lake": "ADD8E6",  # Light blue
    "City": "808080",  # Gray
    "Swamp": "FFD700",  # Gold
    "Desert": "F0E68C",  # Khaki
    "Plain": "7CFC00"  # Dark green
}

# Generate terrain counts
terrain_counts = {
    "Mountain": random.randint(1, 12),
    "Lake": random.randint(0, 4),
    "City": random.randint(2, 4),
    "Swamp": random.randint(0, 3),
    "Desert": random.randint(1, 8)
}

# Generate terrain positions
terrain_positions = {}
for terrain in terrain_counts.keys():
    terrain_positions[terrain] = set()
    while len(terrain_positions[terrain]) < terrain_counts[terrain]:
        x = random.randint(1, 6)
        y = random.randint(1, 6)
        if (x, y) not in terrain_positions[terrain] and \
           all((x, y) not in terrain_positions[t] for t in terrain_positions.keys() if t != terrain):
            terrain_positions[terrain].add((x, y))

# Fill in the worksheet with terrain
for i in range(1, 7):
    for j in range(1, 7):
        for terrain, positions in terrain_positions.items():
            if (i, j) in positions:
                sheet.cell(row=i, column=j).fill = openpyxl.styles.PatternFill(start_color=terrain_colors[terrain], end_color=terrain_colors[terrain], fill_type="solid")
                break
wb.save("C:/Users/James S/Desktop/simulation/Map.xlsx")


print("successful")
file_path = r"C:\Users\James S\Desktop\simulation\Map.xlsx"
if os.path.exists(file_path):
    os.startfile(file_path)
else:
    print("File not found.")
